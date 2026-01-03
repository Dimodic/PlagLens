"""XLSX encoder using openpyxl with simple conditional formatting."""
from __future__ import annotations

import io

from ..builders.base import BuilderResult


def to_xlsx(result: BuilderResult) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (result.title or "Sheet1")[:31]
    ws.append(result.columns)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    danger_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF7CC", end_color="FFF7CC", fill_type="solid")

    for row in result.rows:
        ws.append([row.get(c) for c in result.columns])

    flag_by_idx: dict[tuple[int, int], str] = {}
    col_index = {c: i + 1 for i, c in enumerate(result.columns)}
    for f in result.cell_flags:
        try:
            r = int(f.get("row", "0")) + 2  # header offset
            ci = col_index.get(f.get("column", ""), 0)
            level = f.get("level", "warn")
            if ci:
                flag_by_idx[(r, ci)] = level
        except Exception:
            continue

    # Apply conditional formatting from explicit flags + heuristic on similarity/suspicious cols
    for (r, c), level in flag_by_idx.items():
        cell = ws.cell(row=r, column=c)
        cell.fill = danger_fill if level == "danger" else warn_fill

    # Heuristic: similarity >= 0.85, suspicious_count > 0, average_score < 50
    for ci, name in enumerate(result.columns, start=1):
        if name in ("max_similarity", "similarity"):
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=ci).value
                try:
                    if v is not None and float(v) >= 0.85:
                        ws.cell(row=r, column=ci).fill = danger_fill
                except (TypeError, ValueError):
                    pass
        elif name == "suspicious_count":
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=ci).value
                try:
                    if v is not None and int(v) > 0:
                        ws.cell(row=r, column=ci).fill = warn_fill
                except (TypeError, ValueError):
                    pass
        elif name == "average_score":
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=ci).value
                try:
                    if v is not None and float(v) < 50:
                        ws.cell(row=r, column=ci).fill = warn_fill
                except (TypeError, ValueError):
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
